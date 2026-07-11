import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import pytz
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="Generador de Horarios - Zorro", page_icon="🦊", layout="wide")

st.title("🦊 Generador de Horarios de Capacitación - Grupo Zorro")
st.write("Sube el reporte de Excel para organizar automáticamente a tu equipo esta semana.")

archivo_subido = st.file_uploader("📂 Sube tu archivo aquí (cualquier nombre funciona)", type=['xlsx'])

# ==========================================
# 🧠 CEREBRO CACHEADO (Súper Optimizado)
# ==========================================
@st.cache_data(show_spinner=False)
def generar_reporte_excel(df_entrada):
    df = df_entrada.copy()
    
    # --- LIMPIEZA ---
    df['Progreso_clean'] = df['Progreso'].astype(str).str.strip().str.lower()
    df['Nombre Completo'] = df['Nombre(s)'].astype(str) + ' ' + df['Apellido(s)'].astype(str)
    df['Estatus_Original'] = df['Progreso'].astype(str).str.strip()
    
    terminos_pendientes = ['no iniciado', 'no inciado', 'en proceso', 'en progreso', '0%']
    
    total_cursos = len(df)
    cursos_finalizados = len(df[df['Progreso_clean'] == 'finalizado'])
    filtro_pendientes = df['Progreso_clean'].isin(terminos_pendientes)
    
    porcentaje_avance = (cursos_finalizados / total_cursos) * 100 if total_cursos > 0 else 0.0

    # --- TOP 10 ---
    stats = df.groupby('Nombre Completo').agg(
        Pendientes=('Progreso_clean', lambda x: x.isin(terminos_pendientes).sum()),
        Finalizados=('Progreso_clean', lambda x: (x == 'finalizado').sum())
    ).reset_index()
    
    top_10_peor = stats.sort_values(by='Pendientes', ascending=False).head(10)
    top_10_mejor = stats.sort_values(by=['Pendientes', 'Finalizados'], ascending=[True, False]).head(10)

    # --- LÓGICA DE HORARIOS ---
    df_pendientes = df[filtro_pendientes].copy()
    df_pendientes['Curso_Detalle'] = df_pendientes['Nombre curso'].astype(str) + " (" + df_pendientes['Estatus_Original'] + ")"
    
    resumen = df_pendientes.groupby(['Nombre Completo', 'Puesto']).agg(
        Total_Pendientes=('Curso_Detalle', 'count'),
        Nombres_Cursos=('Curso_Detalle', lambda x: ', '.join(x))
    ).reset_index()

    resumen = resumen.sort_values(by='Total_Pendientes', ascending=False)
    cola_colaboradores = resumen.to_dict('records')
    total_colaboradores = len(cola_colaboradores)

    horarios = []
    
    if total_colaboradores > 0:
        def obtener_limite_semanal(puesto):
            p_lower = str(puesto).lower()
            if any(rol in p_lower for rol in ['gerente', 'subgerente', 'comodin', 'comodín', 'administrativa', 'administrativo']):
                return 2  
            elif 'lider' in p_lower or 'líder' in p_lower:
                return 3  
            else:
                return 4  
                
        limites_semanales = {c['Nombre Completo']: obtener_limite_semanal(c['Puesto']) for c in cola_colaboradores}
        sesiones_semana = {c['Nombre Completo']: 0 for c in cola_colaboradores}

        indice_colaborador = 0
        tz_mx = pytz.timezone('America/Mexico_City')
        hoy = datetime.now(tz_mx)
        fecha_actual = hoy + timedelta(days=1)
        
        dias_semana = {0: 'lunes', 1: 'martes', 2: 'miércoles', 3: 'jueves', 4: 'viernes', 5: 'sábado', 6: 'domingo'}
        meses = {1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril', 5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto', 9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'}
        
        dias_programados = 0
        
        while dias_programados < 7:
            hora_actual = fecha_actual.replace(hour=8, minute=0, second=0, microsecond=0)
            hora_limite = fecha_actual.replace(hour=20, minute=0, second=0, microsecond=0)
            inicio_bloqueo = fecha_actual.replace(hour=13, minute=30, second=0, microsecond=0)
            fin_bloqueo = fecha_actual.replace(hour=14, minute=30, second=0, microsecond=0)
            fecha_texto = f"{dias_semana[fecha_actual.weekday()]} {fecha_actual.day} de {meses[fecha_actual.month]}"
            
            sesiones_hoy = {c['Nombre Completo']: 0 for c in cola_colaboradores}
            
            while hora_actual < hora_limite:
                hora_fin = hora_actual + timedelta(minutes=25) 
                
                if hora_actual < fin_bloqueo and hora_fin > inicio_bloqueo:
                    horarios.append({
                        'Fecha y Horario': f"{fecha_texto} de {hora_actual.strftime('%H:%M')} a {fin_bloqueo.strftime('%H:%M')}", 
                        'Colaborador': '⚠️ RECESO / OPERACIÓN', 'Puesto': '---', 
                        'Cursos Pendientes (Detalle)': 'Espacio reservado para operación', 'Total': 0
                    })
                    hora_actual = fin_bloqueo
                    continue 
                
                encontrado = False
                intentos = 0
                while intentos < total_colaboradores:
                    colab = cola_colaboradores[indice_colaborador]
                    nombre = colab['Nombre Completo']
                    
                    if sesiones_semana[nombre] < limites_semanales[nombre] and sesiones_hoy[nombre] < 1:
                        encontrado = True
                        break
                    indice_colaborador = (indice_colaborador + 1) % total_colaboradores
                    intentos += 1
                    
                if encontrado:
                    horarios.append({
                        'Fecha y Horario': f"{fecha_texto} de {hora_actual.strftime('%H:%M')} a {hora_fin.strftime('%H:%M')}",
                        'Colaborador': nombre, 'Puesto': colab['Puesto'],
                        'Cursos Pendientes (Detalle)': colab['Nombres_Cursos'], 'Total': colab['Total_Pendientes']
                    })
                    sesiones_semana[nombre] += 1
                    sesiones_hoy[nombre] += 1
                    indice_colaborador = (indice_colaborador + 1) % total_colaboradores
                else:
                    horarios.append({
                        'Fecha y Horario': f"{fecha_texto} de {hora_actual.strftime('%H:%M')} a {hora_fin.strftime('%H:%M')}",
                        'Colaborador': 'Libre / Sin Asignar', 'Puesto': '---', 
                        'Cursos Pendientes (Detalle)': '---', 'Total': 0
                    })
                hora_actual = hora_fin
            fecha_actual += timedelta(days=1)
            dias_programados += 1

    df_horarios = pd.DataFrame(horarios) if horarios else pd.DataFrame(columns=['Fecha y Horario', 'Colaborador', 'Puesto', 'Cursos Pendientes (Detalle)', 'Total'])
    
    if not df_horarios.empty:
        df_horarios = df_horarios[df_horarios['Colaborador'] != 'Libre / Sin Asignar'].reset_index(drop=True)
        
        # --- NUEVAS REGLAS DE SEGMENTACIÓN EXTENDIDAS ---
        def categorizar_puesto(puesto):
            p_lower = str(puesto).lower()
            if p_lower == '---': 
                return 'Z_Recesos' 
            elif any(rol in p_lower for rol in ['gerente', 'subgerente', 'comodin', 'comodín', 'administrativa', 'administrativo']): 
                return 'A_Gerencia'
            elif 'lider' in p_lower or 'líder' in p_lower: 
                return 'B_Lideres'
            # 🆕 REGLA NOCTURNOS: Cualquier puesto que diga nocturno
            elif 'nocturno' in p_lower:
                return 'C_Nocturnos'
            elif 'exprezo' in p_lower: 
                return 'D_Exprezo'
            elif 'caja' in p_lower or 'cajer' in p_lower: 
                return 'E_Cajas'
            elif 'cremeria' in p_lower or 'cremería' in p_lower: 
                return 'F_Cremeria'
            # 🆕 REGLA RECIBO: Surtidor, Bodega, Recibo, Chofer
            elif any(k in p_lower for k in ['surtidor', 'bodega', 'recibo', 'chofer']):
                return 'G_Recibo'
            # Autoservicio ahora conserva perecederos y farmacia (Surtidores se movieron arriba a recibo)
            elif any(k in p_lower for k in ['autoservicio', 'perecedero', 'farmacia']): 
                return 'H_Autoservicio'
            else: 
                return 'I_Otros Puestos'

        df_horarios['Categoria_Orden'] = df_horarios['Puesto'].apply(categorizar_puesto)
        df_horarios = df_horarios.sort_values(by=['Categoria_Orden', 'Colaborador', 'Fecha y Horario']).reset_index(drop=True)
        
        rows_with_headers = []
        ultima_categoria = None
        
        nombres_areas = {
            'A_Gerencia': '💼 PERSONAL DE GERENCIA Y ADMINISTRACIÓN',
            'B_Lideres': '⭐ EQUIPO DE LÍDERES Y SUBLÍDERES',
            'C_Nocturnos': '🌙 EQUIPO DE PERSONAL NOCTURNO',
            'D_Exprezo': '☕ EQUIPO DE EXPREZO',
            'E_Cajas': '💳 EQUIPO DE CAJAS',
            'F_Cremeria': '🧀 EQUIPO DE CREMERÍA',
            'G_Recibo': '📦 PUESTOS DE RECIBO (RECIBO, BODEGAS, SURTIDORES Y CHOFERES)',
            'H_Autoservicio': '🛒 EQUIPO DE AUTOSERVICIO, PERECEDEROS Y FARMACIA',
            'I_Otros Puestos': '🔧 OTROS PUESTOS OPERATIVOS',
            'Z_Recesos': '⚠️ RECESOS Y BLOQUEOS OPERATIVOS EN TIENDA'
        }
        
        for idx, row in df_horarios.iterrows():
            cat = row['Categoria_Orden']
            if cat != ultima_categoria:
                if ultima_categoria is not None:
                    rows_with_headers.append({'Es_Blanco': True, 'Es_Header': False})
                rows_with_headers.append({
                    'Es_Blanco': False, 'Es_Header': True, 
                    'Titulo': nombres_areas.get(cat, 'ÁREA OPERATIVA'), 'Categoria_Orden': cat
                })
                ultima_categoria = cat
            
            data_row = row.to_dict()
            data_row['Es_Blanco'] = False
            data_row['Es_Header'] = False
            rows_with_headers.append(data_row)
    else:
        rows_with_headers = []

    # --- GENERAR EXCEL EN MEMORIA ---
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Horarios'
    
    borde = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'), top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))
    font_blanca = Font(color="FFFFFF", bold=True, size=11)
    font_negra = Font(color="000000", size=10)
    font_gris = Font(color="808080", size=10)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", indent=1)
    align_wrap = Alignment(vertical="center", wrap_text=True)
    align_valign = Alignment(vertical="center")

    fills = {
        'A_Gerencia': PatternFill(start_color="F2F0F7", end_color="F2F0F7", fill_type="solid"),
        'B_Lideres': PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        'C_Nocturnos': PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"), # Gris Noche
        'D_Exprezo': PatternFill(start_color="EAF1F5", end_color="EAF1F5", fill_type="solid"),
        'E_Cajas': PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid"),
        'F_Cremeria': PatternFill(start_color="FFFEE6", end_color="FFFEE6", fill_type="solid"),
        'G_Recibo': PatternFill(start_color="EAD1DC", end_color="EAD1DC", fill_type="solid"), # Rosa Pastel Viejo
        'H_Autoservicio': PatternFill(start_color="EAF2E8", end_color="EAF2E8", fill_type="solid"),
        'I_Otros Puestos': PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid"),
        'Z_Recesos': PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    }

    fill_headers = {
        'A_Gerencia': PatternFill(start_color="44245E", end_color="44245E", fill_type="solid"),
        'B_Lideres': PatternFill(start_color="C65911", end_color="C65911", fill_type="solid"),
        'C_Nocturnos': PatternFill(start_color="262626", end_color="262626", fill_type="solid"), # Banner Oscuro Nocturno
        'D_Exprezo': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
        'E_Cajas': PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
        'F_Cremeria': PatternFill(start_color="7F6000", end_color="7F6000", fill_type="solid"),
        'G_Recibo': PatternFill(start_color="741B47", end_color="741B47", fill_type="solid"), # Banner Recibo Oscuro
        'H_Autoservicio': PatternFill(start_color="385723", end_color="385723", fill_type="solid"),
        'I_Otros Puestos': PatternFill(start_color="595959", end_color="595959", fill_type="solid"),
        'Z_Recesos': PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    }
    
    color_hex = "00cc66" if porcentaje_avance >= 85 else ("ffcc00" if porcentaje_avance >= 50 else "ff4d4d")
    
    worksheet.merge_cells('A1:E2')
    celda_titulo = worksheet['A1']
    celda_titulo.value = f"REPORTE SEMANAL DE CAPACITACIÓN | AVANCE DE SUCURSAL: {porcentaje_avance:.1f}%"
    celda_titulo.font = Font(color="FFFFFF" if porcentaje_avance < 50 or porcentaje_avance >= 85 else "000000", bold=True, size=16)
    celda_titulo.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    celda_titulo.alignment = align_center
    
    for r_t in range(1, 3):
        for c_t in range(1, 6):
            worksheet.cell(row=r_t, column=c_t).border = borde

    headers = ['Fecha y Horario', 'Colaborador', 'Puesto', 'Cursos Pendientes (Detalle)', 'Total']
    header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    for col_num, h_text in enumerate(headers, 1):
        cell = worksheet.cell(row=4, column=col_num, value=h_text)
        cell.fill = header_fill
        cell.font = font_blanca
        cell.alignment = align_center
        cell.border = borde
    worksheet.row_dimensions[4].height = 28

    current_row = 5
    for r in rows_with_headers:
        worksheet.row_dimensions[current_row].height = 28
        
        if r['Es_Blanco']:
            current_row += 1
            continue
            
        if r['Es_Header']:
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            cell = worksheet.cell(row=current_row, column=1, value=r['Titulo'])
            cell.font = font_blanca
            cell.alignment = align_left
            
            cat_header = r.get('Categoria_Orden', 'I_Otros Puestos')
            h_fill = fill_headers.get(cat_header, fill_headers['I_Otros Puestos'])
            
            for c_num in range(1, 6):
                cell_b = worksheet.cell(row=current_row, column=c_num)
                cell_b.fill = h_fill
                cell_b.border = borde
            current_row += 1
            continue
        
        vals = [r['Fecha y Horario'], r['Colaborador'], r['Puesto'], r['Cursos Pendientes (Detalle)'], r['Total']]
        cat = r['Categoria_Orden']
        
        if r['Colaborador'] == '⚠️ RECESO / OPERACIÓN':
            r_fill = fills['Z_Recesos']
            f_actual = font_gris
        else:
            r_fill = fills.get(cat, fills['I_Otros Puestos'])
            f_actual = font_negra
            
        for col_num, val in enumerate(vals, 1):
            cell = worksheet.cell(row=current_row, column=col_num, value=val)
            cell.fill = r_fill
            cell.font = f_actual
            cell.border = borde
            cell.alignment = align_wrap if col_num == 4 else align_valign
            
            if col_num == 5 and isinstance(val, (int, float)) and val > 0:
                if val >= 10:
                    cell.fill = PatternFill(start_color="FF4D4D", end_color="FF4D4D", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif 4 <= val <= 9:
                    cell.fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
                    cell.font = Font(color="000000", bold=True)
                elif 1 <= val <= 3:
                    cell.fill = PatternFill(start_color="00CC66", end_color="00CC66", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = align_center
                
        current_row += 1

    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 32
    worksheet.column_dimensions['C'].width = 24
    worksheet.column_dimensions['D'].width = 46
    worksheet.column_dimensions['E'].width = 12
    worksheet.column_dimensions['F'].width = 4
    worksheet.column_dimensions['G'].width = 34
    worksheet.column_dimensions['H'].width = 24
    
    # --- TABLAS LATERALES ---
    worksheet.merge_cells('G4:H4')
    t_peor = worksheet['G4']
    t_peor.value = "⚠️ TOP 10 - MAYOR REZAGO"
    t_peor.fill = PatternFill(start_color="ff4d4d", end_color="ff4d4d", fill_type="solid")
    t_peor.font = Font(color="FFFFFF", bold=True)
    t_peor.alignment = align_center
    worksheet.cell(row=4, column=7).border = borde
    worksheet.cell(row=4, column=8).border = borde
    
    fill_p = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") 
    font_p = Font(color="9C0006") 
    for i, (idx, fila) in enumerate(top_10_peor.iterrows(), start=5):
        c_n = worksheet.cell(row=i, column=7, value=fila['Nombre Completo'])
        c_c = worksheet.cell(row=i, column=8, value=f"{fila['Pendientes']} cursos por hacer")
        for cell in [c_n, c_c]:
            cell.fill = fill_p; cell.font = font_p; cell.border = borde; cell.alignment = align_valign
        c_c.alignment = align_center

    worksheet.merge_cells('G18:H18')
    t_mejor = worksheet['G18']
    t_mejor.value = "🌟 TOP 10 - MEJOR APROVECHAMIENTO"
    t_mejor.fill = PatternFill(start_color="00cc66", end_color="00cc66", fill_type="solid")
    t_mejor.font = Font(color="FFFFFF", bold=True)
    t_mejor.alignment = align_center
    worksheet.cell(row=18, column=7).border = borde
    worksheet.cell(row=18, column=8).border = borde
    
    fill_m = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") 
    font_m = Font(color="006100") 
    for i, (idx, fila) in enumerate(top_10_mejor.iterrows(), start=19):
        c_n2 = worksheet.cell(row=i, column=7, value=fila['Nombre Completo'])
        c_c2 = worksheet.cell(row=i, column=8, value=f"{fila['Pendientes']} cursos por hacer")
        for cell in [c_n2, c_c2]:
            cell.fill = fill_m; cell.font = font_m; cell.border = borde; cell.alignment = align_valign
        c_c2.alignment = align_center
        
    for r in range(4, 29):
        worksheet.row_dimensions[r].height = 28

    workbook.save(output)
    return output.getvalue(), porcentaje_avance, total_colaboradores

# ==========================================
# 🚀 INTERFAZ PRINCIPAL 
# ==========================================
if archivo_subido is not None:
    try:
        archivo_subido.seek(0) 
        df_crudo = pd.read_excel(archivo_subido)
        
        with st.spinner("🧠 Calculando agenda y armando el Excel con IA..."):
            excel_bytes, porcentaje, total_gente = generar_reporte_excel(df_crudo)
        
        st.subheader("📈 Estado de Capacitación de la Sucursal")
        if porcentaje >= 85:
            st.success(f"¡Excelente ritmo! Avance de la Sucursal: **{porcentaje:.1f}%**")
        elif 50 <= porcentaje < 85:
            st.warning(f"Buen esfuerzo. Avance de la Sucursal: **{porcentaje:.1f}%**")
        else:
            st.error(f"Atención requerida. Avance de la Sucursal: **{porcentaje:.1f}%**")
            
        st.divider()

        if total_gente == 0:
            st.success("🎉 ¡Felicidades! Todo el personal está al 100%.")
        else:
            st.success("✅ ¡Horario premium generado y cacheado con éxito!")
            
            st.download_button(
                label="📥 Descargar Horario Final", 
                data=excel_bytes, 
                file_name="Horarios_Zorro_Premium.xlsx", 
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"❌ Error al procesar el archivo: Verifica que sea el Excel correcto. (Detalle: {e})")